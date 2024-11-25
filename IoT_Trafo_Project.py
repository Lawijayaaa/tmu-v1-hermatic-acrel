#!/usr/bin/env python3
import openpyxl, requests
import time, datetime, sys, os
import mysql.connector, math, logging, threading
from toolboxTmu import *
from pymodbus.client import ModbusSerialClient
from requests.models import StreamConsumedError
from requests.exceptions import Timeout
from tkinter import *
from openpyxl import *

os.chdir('/home/pi/tmu-v1-hermatic/')

#init value
engineName = "Trafo X"
teleURL = 'http://192.168.4.120:1444/api/transformer/sendNotificationToTelegramGroup'
reverser = 1
limitPF = 3
healthLimit = 90
trialNumLimit = 2
nCounter = 17
loadCoef = 60   #Wti change direction
cycleSpan = 10   #60 sec / 6 sec/cycle
designedKrated = 1  #initial value
eddyLosesGroup = 0.02 #See table Eddy Current Group
progStat = True
CTratio = 1
PTratio = 1
CTPTratio = PTratio * CTratio

#init tkinter
screen = Tk()
screen.title("IoT Trafo Gateway")
window_width = 400
window_height = 225
screen_width = screen.winfo_screenwidth()
screen_height = screen.winfo_screenheight()
center_x = int(screen_width/2 - window_width/2)
center_y = int(screen_height/2 - window_height/2)
screen.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
screen.resizable(0,0)
screen.attributes('-topmost', 1)
screen.configure(background = "#17C0EB")

#init logger
ts = time.strftime("%Y%m%d")
logName = r'/home/pi/tmu/tmu-app-client-deploy/assets/datalog/sysdata/syslog-' + ts + '.log'
logging.basicConfig(filename=logName, format='(asctime)s | %(levelname)s: %(message)s',level=logging.INFO)
pathDatLog = r'/home/pi/tmu/tmu-app-client-deploy/assets/datalog/rawdata/datalogger-' + ts + '.xlsx'
sheetName = ["Harmonic_phR","Harmonic_phS", "Harmonic_phT"]

#create datalog
try:
    wb = openpyxl.load_workbook(pathDatLog)
    logging.info("Open Existing Datalog")
except:
    logging.exception("Error Opening Datalog")
    workbook = Workbook()
    for member in sheetName:
        workbook.create_sheet(member)
    workbook.save(pathDatLog)
    logging.info("Create New Datalog")
    wb = openpyxl.load_workbook(pathDatLog)
    sheet = wb.active
    sheet.title = "Raw_data"
    name = (('timestamp', 'OilTemp', 'BusTemp1', 'BusTemp2', 'BusTemp3', 'Press', 'Level',
            'Van', 'Vab', 'Ia', 'PFa', 'Pa', 'Qa', 'Sa', 'THDV1', 'THDI1',
            'Vbn', 'Vbc', 'Ib', 'PFb', 'Pb', 'Qb', 'Sb', 'THDV2', 'THDI2',
            'Vcn', 'Vca', 'Ic', 'PFc', 'Pc', 'Qc', 'Sc', 'THDV3', 'THDI3',
            'Itot', 'PFsyst', 'Psig', 'Qsig', 'Ssig', 'Freq', 'Ineutral',
            'kWhInp', 'kWhOut', 'kVARhinp', 'kVARhOut',
            'KRateda', 'deRatinga', 'KRatedb', 'deRatingb', 'KRatedc', 'deRatingc',
            'WTITemp1', 'WTITemp2', 'WTITemp3', 
            'trafoStat', 'DIstat', 'DOstat', 'Alarm', 'Trip1', 'Trip2'),)
    for row in name:
        sheet.append(row)
    wb.save(pathDatLog)
    for name in sheetName:
        sheetHarm = wb[name]
        name = (('timestamp', 'V 1st', 'V 3rd' , 'V 5th' , 'V 7th' , 'V 9th' , 'V 11th' , 'V 13th' , 'V 15th' ,
                 'V 17th' , 'V 19th' , 'V 21st' , 'V 23rd' , 'V 25th' , 'V 27th' , 'V 29th' , 'V 31st',
                 'I 1st', 'I 3rd' , 'I 5th' , 'I 7th' , 'I 9th' , 'I 11th' , 'I 13th' , 'I 15th' ,
                 'I 17th' , 'I 19th' , 'I 21st' , 'I 23rd' , 'I 25th' , 'I 27th' , 'I 29th' , 'I 31st'),)
        for row in name:
            sheetHarm.append(row)
        wb.save(pathDatLog)

#init modbus device, db
client = ModbusSerialClient(method='rtu', port='/dev/ttyACM0', baudrate=9600)
db = mysql.connector.connect(
    host = "localhost",
    user = "client",
    passwd = "raspi",
    database= "iot_trafo_client")
    
#fetch init trafoData
cursor0 = db.cursor()
sql0 = "SELECT * FROM transformer_data"
cursor0.execute(sql0)
initData = cursor0.fetchall()[0]
suddenState = initData[29]

#pre-init used Var    
windTemp = [29.0, 28.0, 31.0]
timePassed = [1, 1, 1]
deltaHi1 = [0, 0 , 0]
deltaHi2 = [0, 0, 0]
lastLoadDefiner = [0, 0.0, 0.0]
currentLoadDefiner = [0.0, 0.0, 0.0]
deltaH1 = [0, 0, 0]
deltaH2 = [0, 0, 0]
raisingLoadBool = [True, True, True]
loadFactor = [0.0, 0.0, 0.0]
trialNum = 0
gasFault = False
counterPF = 0
DIstat = [0]*9
sampleNum = 0 #GPS

def plcHandler(getPLC):
    plcData = [0]*5
    try:
        plcData[0:4] = getPLC.registers
        plcData[4] = 500
    except:
        plcData[4] = 404
        pass
    if plcData[0] < 194:
        plcData[0] = 0
    else:
        plcData[0] = (round(((plcData[0] - 192.324)/769.296)*100))/100 #Pressure Calibration
    return plcData

def dataHandler(getTemp, getOil, getElect1, getElect2, getElect3, getHarmV, getHarmA, currentResult, CTratio, PTratio):
    try:
        currentResult[0] = (round(((0.195 * getOil.registers[0]) - 37.5)*100))/100 #oiltemp
        currentResult[1:4] = [member/10 for member in getTemp.registers] #bustemp
        for i in range(1, 4):
            if currentResult[i]>240:
                currentResult[i] = 0
        print(currentResult[0:4])
    except:
        pass
    try:
        for i in range(0,3):
            currentResult[(i*9)+6] = (PTratio * getElect1.registers[i])/100
            currentResult[(i*9)+7] = (PTratio * getElect1.registers[i+3])/100
            currentResult[(i*9)+8] = (CTratio * getElect1.registers[i+6])/1000

            currentResult[(i*9)+9] = (signedInt16Handler(getElect1.registers[i+21]))/1000
            currentResult[(i*9)+10] = (CTPTratio * signedInt16Handler(getElect1.registers[i+15]))/10
            currentResult[(i*9)+11] = (CTPTratio * signedInt16Handler(getElect1.registers[i+18]))/10

        currentResult[39] = (CTratio * getElect1.registers[9])/1000
        currentResult[35] = ((CTPTratio * signedInt32Handler(getElect1.registers[10:12]))[0])/10
        currentResult[36] = ((CTPTratio * signedInt32Handler(getElect1.registers[12:14]))[0])/10
        currentResult[34] = (getElect1.registers[14])/1000
        currentResult[38] = (getElect1.registers[24])/100
        currentResult[40] = (unsignedInt32Handler(getElect1.registers[25:27]))/10
        currentResult[42] = (unsignedInt32Handler(getElect1.registers[27:]))/10
        currentResult[41] = currentResult[43] = 0
        currentResult[33] = currentResult[8] + currentResult[17] +  currentResult[26]
    except:
        pass
    try:
        for i in range(0, 3):
            currentResult[(i*9)+12] = CTPTratio * (getElect2.registers[i])/10
        currentResult[37] = (unsignedInt32Handler(getElect2.registers[3:]))/10
    except:
        pass
    try:
        for i in range(0, 3):
            currentResult[(i*9) + 13] = (getElect3.registers[i])/10
            currentResult[(i*9) + 14] = (getElect3.registers[i+3])/10
    except:
        pass
    try:
        Vharm = [[0]*15, [0]*15, [0]*15]
        Iharm = [[0]*15, [0]*15, [0]*15]
        harmV = [[0]*15, [0]*15, [0]*15]
        harmA = [[0]*15, [0]*15, [0]*15]

        harmV[0] = getHarmV.registers[0:30]
        harmV[1] = getHarmV.registers[30:60]
        harmV[2] = getHarmV.registers[60:]
        
        harmA[0] = getHarmA.registers[0:30]
        harmA[1] = getHarmA.registers[30:60]
        harmA[2] = getHarmA.registers[60:]
        harmIndex = 0
        for i in range(0, 3):
            for j in range(0, len(harmV[i])):
                if j % 2 == 1 :
                    Vharm[i][harmIndex] = (harmV[i][j])/10
                    Iharm[i][harmIndex] = (harmA[i][j])/10
                    harmIndex = harmIndex + 1
            harmIndex = 0
            Vharm[i].insert(0, 100)
            Iharm[i].insert(0, 100)
    except:
        Vharm = [[0]*16, [0]*16, [0]*16]
        Iharm = [[0]*16, [0]*16, [0]*16]
        pass

    return currentResult, Vharm, Iharm

def Restart():
    logging.info("Saving Excel File before Restart")
    global wb
    wb.save(pathDatLog)
    logging.info("Restart")
    os.execv(sys.executable, [sys.executable] + ['/home/pi/tmu-v1-hermatic/IoT_Trafo_Project.py'])

def Start():
    global progStat
    progStat = True

def Stop():
    global progStat
    progStat = False
    
def mainLoop(thread_name, interval):
    logging.info("Program Started")
    global windTemp, timePassed, deltaHi1, deltaHi2, lastLoadDefiner, currentLoadDefiner, deltaH1, deltaH2, counterPF 
    global raisingLoadBool, loadFactor, trialNum, gasFault, DIstat, wb, pathDatLog, sheetName, progStat, lastPosition
    paramName = ['Voltage a-n ', 'Voltage b-n ', 'Voltage c-n ', 'Voltage a-b ', 'Voltage b-c ', 'Voltage a-c ',
                'Difference Between Van - Vbn ', 'Difference Between Vbn - Vcn ', 'Difference Between Van - Vcn ', 
                'Frequency ', 'Top Oil Temperature ', 'Winding Phase a Temperature ', 'Winding Phase b Temperature ', 
                'Winding Phase c Temperature ', 'Busbar Phase a Temperature ', 'Busbar Phase b Temperature ', 'Busbar Phase c Temperature ',  
                'Power Factor ', 'LV Current a ', 'LV Current b ', 'LV Current c ', 'Ambient Temperature ', 'Pressure ', 'Oil Level ', 
                'THD Current a', 'THD Current b', 'THD Current c', 'THD Voltage a', 'THD Voltage b', 'THD Voltage c', 'Neutral Current']
    messageReason = ['Extreme Low',
                'Low', 
                'Back Normal', 
                'High', 
                'Extreme High']
    anchorDay = datetime.datetime.now().day
    cursor6 = db.cursor()
    sql6 = "SELECT * FROM failure_log"
    cursor6.execute(sql6)
    listFailure = cursor6.fetchall()
    activeFailure = []
    for i in range(0, len(listFailure)):
        if listFailure[i][2] == None :
            activeFailure.append(listFailure[i])
    logging.info("Last Active Failure : " + str(activeFailure))
    db.commit()
    currentResult = [0]*53
    newResult = [[0]*53, [[0.0] * 16] * 3, [[0.0] * 16] * 3]
    paramNum = len(paramName)
    stateListBool = [0]*paramNum
    errorMsg = [""]*paramNum
    errorMsgReg = [""]*paramNum
    previousTime = excelPrevTime = datetime.datetime.now()
    while True:
        while progStat:
            logging.info("Start Sect. Looping")
            progStatLbl['text'] = "Running"
            currentTime = datetime.datetime.now()
            if int((currentTime-previousTime).total_seconds()) > 3600:
                logging.info("D02 Hourly Fetch Active Error")
                for i in range(0, len(activeFailure)):
                    errorMsgReg[i] = str(activeFailure[i][4] + " " + activeFailure[i][3] + " , Value = " + activeFailure[i][5] + "\n" + "Time Occurence : " + str(activeFailure[i][1]))
                errorMsgSendReg = list(filter(None, errorMsgReg))
                if activeFailure :
                    logging.info("D03 Hourly Remind if Any Fault")
                    sendTelegramReg = "\n-----------\n".join(map(str, errorMsgSendReg))
                    messages = engineName + " Says : " + "\n" + sendTelegramReg
                    ploadReg = {'message':messages}
                    try:
                        r = requests.post(teleURL, data = ploadReg, timeout = 5, verify = False)
                    except Timeout:
                        logging.warning("Timeout while sending tele message")
                    except Exception as Argument:
                        logging.exception("Tele Catch")
                else:
                    logging.info("D04 No alarm/trip within an hours")
                previousTime = datetime.datetime.now()
            thisDay = currentTime.day
            if thisDay != anchorDay:
                logging.info("Saving Excel File")
                wb.save(pathDatLog)
                logging.info("Excel File Saved")
                Restart()
            logging.info("D05 Fetching init data from db")
            cursor3 = db.cursor()
            sql3 = "SELECT * FROM transformer_settings"
            cursor3.execute(sql3)
            trafoSetting = cursor3.fetchall()[0] 
            cursor2 = db.cursor()
            sql2 = "SELECT * FROM transformer_data"
            cursor2.execute(sql2)
            trafoData = cursor2.fetchall()[0]
            hotspotFactor = trafoData[25]
            gradient = trafoData[21]
            coolingMode = trafoData[27]
            cursor14 = db.cursor()
            sql14 = 'SELECT WindingExponent, K21, K22, T0, Tw FROM constanta_value WHERE typeCooling = %s '
            valAssign = (str(coolingMode), )
            cursor14.execute(sql14, valAssign)
            constantWTI = cursor14.fetchall()[0]
            cursor15 = db.cursor()
            sql15 = 'SELECT number, name, state FROM di_scan'
            cursor15.execute(sql15)
            DIprop = cursor15.fetchall()
            db.commit()
            stateChange = False
            logging.info("D06 creating Active Param List")
            activeParam = [0]*paramNum
            if len(activeFailure) > 0:
                for k in range(0, len(activeFailure)):
                    activeParam[k] = activeFailure[k][4]
            else:
                activeParam[0] = None     
            logging.info("D07 get data from ModBus Devices")
            getTemp = client.read_holding_registers(0, 3, slave = 3)
            getPLC = client.read_holding_registers(55, 4, slave = 1)
            getOil = client.read_holding_registers(54, 1, slave = 1)
            getElect1 = client.read_holding_registers(0, 29, slave = 2)
            getElect2 = client.read_holding_registers(46, 5, slave = 2)
            getElect3 = client.read_holding_registers(800, 6, slave = 2)
            getHarmV = client.read_holding_registers(806, 90, slave = 2)
            getHarmA = client.read_holding_registers(896, 90, slave = 2)
            logging.info("D08 Handling received data")
            plcResult = [0]*5
            try:
                newResult = dataHandler(getTemp, getOil, getElect1, getElect2, getElect3, getHarmV, getHarmA, currentResult, CTratio, PTratio)
                plcResult = plcHandler(getPLC)
                if plcResult[4] == 500:
                    newResult[0][4:6] = plcResult[0:2]
            except Exception as Argument:
                logging.info(newResult)
                try:
                    newResult[0] = [0]*53
                    for i in range(0,3):
                        newResult[1][i] = [0]*16
                        newResult[2][i] = [0]*16
                except Exception as Argument:
                    logging.exception("Couldn't create new empty Result")
                logging.exception("Data Handling Error")
            binData = format(plcResult[3], "b")
            leftOver = 9 - len(binData)
            addition = ""
            if leftOver :
                for i in range(0, leftOver): addition = addition + '0'
                binData = addition + binData
            for i in range(0, len(binData)): DIstat[i] = binData[i]
            DIstat.reverse()
            logging.info("CPU Load : " + str(os.getloadavg()))
            logging.info("D09 Process0 : Check data healthness")
            healthResult = checkDataHealthness(newResult[0][0:44], healthLimit)
            if healthResult == True or trialNum >= trialNumLimit-1:
                logging.info("D10 Data health passed")
                trialNum = 0
                gasFault = False
                logging.info("D11 Process1 : Check DI Changes")
                for i in range (0, 9):
                    if str(DIstat[i]) != str(DIprop[i][2]):
                        if str(DIstat[i]) == '0':
                            DIstringStat = " become False"
                        else:
                            DIstringStat = " become True"
                        cursor16 = db.cursor()
                        sql16 = """UPDATE di_scan SET state = %s WHERE number = %s"""
                        cursor16.execute(sql16, (DIstat[i], i))
                        logging.info("Sending Info Message to Telegram")
                        messages = engineName + " Says : " + "\n" + "DI Changed : " + str(DIprop[i][1]) + DIstringStat
                        pload = {'message':messages}
                        try:
                            r = requests.post(teleURL, data = pload, timeout = 5, verify = False)
                        except Timeout:
                            logging.warning("Timeout while sending tele message")
                        except Exception as Argument:
                            logging.exception("Tele Catch")   
                logging.info("D12 Process2 : Calculate WTI")
                for i in range(0, 3): loadFactor[i] = (currentResult[(8*(i+1))+i]/trafoData[6])
                for i in range(0, 3):
                    currentLoadDefiner[i] = currentResult[i*9]
                    if currentLoadDefiner[i] - lastLoadDefiner[i] >= loadCoef:
                        timePassed[i] = 1
                        deltaHi1[i] = deltaH1[i]
                        deltaHi2[i] = deltaH2[i]
                        raisingLoadBool[i] = True
                        lastLoadDefiner[i] = currentLoadDefiner[i]
                    elif lastLoadDefiner[i] - currentLoadDefiner[i] >= loadCoef:
                        timePassed[i] = 1
                        deltaHi1[i] = deltaH1[i]
                        deltaHi2[i] = deltaH2[i]
                        raisingLoadBool[i] = False
                        lastLoadDefiner[i] = currentLoadDefiner[i]
                    else:
                        timePassed[i] = timePassed[i] + 1
                    try:
                        if raisingLoadBool[i]:
                            deltaH1[i] = deltaHi1[i] + (((constantWTI[1]*hotspotFactor*gradient)*(math.pow(loadFactor[i], constantWTI[0])) - deltaHi1[i])*(1 - math.exp(timePassed[i]/((-1 * constantWTI[2] * constantWTI[4])*3))))
                            deltaH2[i] = deltaHi2[i] + ((((constantWTI[1] - 1)*hotspotFactor*gradient)*(math.pow(loadFactor[i], constantWTI[0])) - deltaHi2[i])*(1 - math.exp(timePassed[i]/(((-1 * constantWTI[3])/constantWTI[2])*3))))
                        else:
                            deltaH1[i] = constantWTI[1] * hotspotFactor * gradient * math.pow(loadFactor[i], constantWTI[0]) + (deltaHi1[i] - (constantWTI[1] * hotspotFactor * gradient * math.pow(loadFactor[i], constantWTI[0])))*(math.exp((-1 * timePassed[i])/(constantWTI[2] * constantWTI[4])))
                            deltaH2[i] = (constantWTI[1] - 1) * hotspotFactor * gradient * math.pow(loadFactor[i], constantWTI[0]) + (deltaHi2[i] - (constantWTI[1] - 1) * hotspotFactor * gradient * math.pow(loadFactor[i], constantWTI[0]))*(math.exp((-1 * timePassed[i])/(constantWTI[3]/constantWTI[4])))
                    except Exception as Argument:
                        logging.exception("Error While Calculating")
                    windTemp[i] = currentResult[0] + (deltaH1[i] - deltaH2[i])
                    currentResult[i+50] = (round(windTemp[i]*100))/100
                logging.info("D13 Process3 : Update harmonics db and calculate derating")
                cursor5 = db.cursor()
                sql5 = """UPDATE voltage_harmonic SET 
                    1st = %s , 3rd = %s , 5th = %s , 7th = %s , 9th = %s , 11th = %s , 
                    13th = %s , 15th = %s , 17th = %s , 19th = %s , 21th = %s , 23th = %s , 
                    25th = %s , 27th = %s , 29th = %s , 31th = %s WHERE Phase = 1"""
                cursor5.execute(sql5, newResult[1][0])
                sql5 = """UPDATE voltage_harmonic SET 
                    1st = %s , 3rd = %s , 5th = %s , 7th = %s , 9th = %s , 11th = %s , 
                    13th = %s , 15th = %s , 17th = %s , 19th = %s , 21th = %s , 23th = %s , 
                    25th = %s , 27th = %s , 29th = %s , 31th = %s WHERE Phase = 2"""
                cursor5.execute(sql5, newResult[1][1])
                sql5 = """UPDATE voltage_harmonic SET 
                    1st = %s , 3rd = %s , 5th = %s , 7th = %s , 9th = %s , 11th = %s , 
                    13th = %s , 15th = %s , 17th = %s , 19th = %s , 21th = %s , 23th = %s , 
                    25th = %s , 27th = %s , 29th = %s , 31th = %s WHERE Phase = 3"""
                cursor5.execute(sql5, newResult[1][2])
                cursor5 = db.cursor()
                sql5 = """UPDATE current_harmonic SET 
                    1st = %s , 3rd = %s , 5th = %s , 7th = %s , 9th = %s , 11th = %s , 
                    13th = %s , 15th = %s , 17th = %s , 19th = %s , 21th = %s , 23th = %s , 
                    25th = %s , 27th = %s , 29th = %s , 31th = %s WHERE Phase = 1"""
                cursor5.execute(sql5, newResult[2][0])
                sql5 = """UPDATE current_harmonic SET 
                    1st = %s , 3rd = %s , 5th = %s , 7th = %s , 9th = %s , 11th = %s , 
                    13th = %s , 15th = %s , 17th = %s , 19th = %s , 21th = %s , 23th = %s , 
                    25th = %s , 27th = %s , 29th = %s , 31th = %s WHERE Phase = 2"""
                cursor5.execute(sql5, newResult[2][1])
                sql5 = """UPDATE current_harmonic SET 
                    1st = %s , 3rd = %s , 5th = %s , 7th = %s , 9th = %s , 11th = %s , 
                    13th = %s , 15th = %s , 17th = %s , 19th = %s , 21th = %s , 23th = %s , 
                    25th = %s , 27th = %s , 29th = %s , 31th = %s WHERE Phase = 3"""
                cursor5.execute(sql5, newResult[2][2])
                for i in range(0, 3):
                    sendHarm = [datetime.datetime.now().strftime("%H:%M:%S")] + newResult[1][i] + newResult[2][i]
                    sendHarm = ((tuple(sendHarm)),)
                    sheetHarm = wb[sheetName[i]]
                    for row in sendHarm:
                        sheetHarm.append(row)
                kFactor = [0, 0, 0]
                kRated = [0, 0, 0]
                deRating = [0, 0, 0]
                kRatedList = newResult[2]
                nHarmonic = 32
                hSquared = [0]*nHarmonic
                for i in range(0, nHarmonic):
                    hSquared[i] = math.pow(((2*(i+1))-1), 2)
                for i in range(0, len(kRatedList)):
                    for j in range(0, len(kRatedList[i])):
                        kRatedList[i][j] = math.pow((newResult[2][i][j])/100, 2) * hSquared[j]
                    kFactor[i] = sum(kRatedList[i])
                    kRated[i] = round(kFactor[i])
                    deRating[i] = 100 * (math.pow((eddyLosesGroup + 1)/(kRated[i]*eddyLosesGroup + 1), 0.8) - math.pow((eddyLosesGroup + 1)/(designedKrated*eddyLosesGroup + 1), 0.8) + 1)
                    if deRating[i] > 100 :
                        deRating[i] = 100
                    else:
                        deRating[i] = deRating[i]  
                    currentResult[i*2 + 44] = kRated[i]
                    currentResult[i*2 + 45] = (round(deRating[i] * 100))/100
                logging.info("D14 Process4 : Insert new data to db")
                cursor = db.cursor()
                sql = """INSERT INTO reading_data (timestamp,
                                            OilTemp, BusTemp1, BusTemp2, BusTemp3, Press, Level,
                                            Van, Vab, Ia, PFa, Pa, Qa, Sa, THDV1, THDI1,
                                            Vbn, Vbc, Ib, PFb, Pb, Qb, Sb, THDV2, THDI2,
                                            Vcn, Vca, Ic, PFc, Pc, Qc, Sc, THDV3, THDI3,
                                            Iavg, PFsig, Psig, Qsig, Ssig, Freq, Ineutral,
                                            kWhInp, kWhOut, kVARhinp, kVARhOut, KRateda, deRatinga,
                                            KRatedb, deRatingb, KRatedc, deRatingc,
                                            WTITemp1, WTITemp2, WTITemp3) VALUES
                                            (%s, %s, %s, %s, %s, %s, %s,
                                            %s, %s, %s, %s, %s, %s, %s,
                                            %s, %s, %s, %s, %s, %s, %s,
                                            %s, %s, %s, %s, %s, %s, %s,
                                            %s, %s, %s, %s, %s, %s, %s,
                                            %s, %s, %s, %s, %s, %s, %s,
                                            %s, %s, %s, %s, %s, %s, %s,
                                            %s, %s, %s, %s, %s)"""
                sendData = [datetime.datetime.now()] + currentResult
                cursor.execute(sql, sendData)
                db.commit()
                logging.info("New data added to db")
                lastTsLbl['text'] = str(datetime.datetime.now())
                logging.info("D15 Process5 : Conditioning Setting and Value")
                [paramValue, gasFault] = gatherParamValue(currentResult, paramNum, gasFault)
                paramThreshold = gatherSetting(trafoSetting, trafoData, paramNum)
                for i in range(0, 3):
                    paramThreshold[2][i + 18] = (paramThreshold[2][i + 18] * deRating[i])/100
                    paramThreshold[3][i + 18] = (paramThreshold[3][i + 18] * deRating[i])/100                
                paramPrint = [0]*paramNum
                for i in range(0, paramNum):
                    if i == 18 or i == 19 or i == 20:
                        paramPrint[i] = (paramValue[i] / trafoData[6]) * 100
                        paramPrint[i] = str(paramPrint[i]) + " Percent , Rated Current = " + str(trafoData[6])
                    else:
                        paramPrint[i] = paramValue[i]      
                logging.info("D15 Process5 : Deciding new Trafo Status")
                cursor4 = db.cursor()
                sql4 = "SELECT * FROM transformer_status"
                cursor4.execute(sql4)
                prevState = cursor4.fetchall()[0][1:paramNum+1]
                currentState = [0]*paramNum
                previousState = [0]*paramNum
                for i in range(0, paramNum):
                    previousState[i] = prevState[i]
                    if paramThreshold[0][i] > 0 and paramThreshold[2][i] > 0:
                        if paramValue[i] <= paramThreshold[0][i]:
                            currentState[i] = 1
                        elif paramValue[i] > paramThreshold[0][i] and paramValue[i] <= paramThreshold[1][i]:
                            currentState[i] = 2
                        elif paramValue[i] > paramThreshold[1][i] and paramValue[i] < paramThreshold[2][i]:
                            currentState[i] = 3
                        elif paramValue[i] >= paramThreshold[2][i] and paramValue[i] < paramThreshold[3][i]:
                            currentState[i] = 4
                        elif paramValue[i] >= paramThreshold[3][i]:
                            currentState[i] = 5
                    elif paramThreshold[0][i] < 1 and paramThreshold[2][i] > 0:
                        if paramValue[i] < paramThreshold[2][i]:
                            currentState[i] = 3
                        elif paramValue[i] >= paramThreshold[2][i] and paramValue[i] < paramThreshold[3][i]:
                            currentState[i] = 4
                        elif paramValue[i] >= paramThreshold[3][i]:
                            currentState[i] = 5
                    elif paramThreshold[0][i] > 0 and paramThreshold[2][i] < 1:
                        if paramValue[i] <= paramThreshold[0][i]:
                            currentState[i] = 1
                        elif paramValue[i] > paramThreshold[0][i] and paramValue[i] <= paramThreshold[1][i]:
                            currentState[i] = 2
                        elif paramValue[i] > paramThreshold[1][i]:
                            currentState[i] = 3            
                if currentState[17] < 3:
                    if currentResult[33] == 0:
                        currentState[17] = 3
                    elif currentResult[36] < 0:
                        currentState[17] = 3
                    elif counterPF < limitPF:
                        currentState[17] == 3
                        counterPF = counterPF + 1
                    elif counterPF == limitPF:
                        counterPF = 0
                        currentState[17] = currentState[17]
                if max(currentState) == 3 and min(currentState) == 3:
                    trafoStateStr = "Safe"
                elif max(currentState) == 5 or min(currentState) == 1:
                    trafoStateStr = "Trip"
                elif max(currentState) == 4 or min(currentState) == 2:
                    trafoStateStr = "Alarm"
                sendLog = sendData[1:]
                sendLog = [datetime.datetime.now().strftime("%H:%M:%S")] + sendLog + [trafoStateStr]
                sendLog = ((tuple(sendLog)),)
                sheet = wb["Raw_data"]
                for row in sendLog:
                    sheet.append(row)
                excelCurrTime = datetime.datetime.now()
                if int((excelCurrTime-excelPrevTime).total_seconds()) > 300:
                    logging.info("Saving Excel File")
                    wb.save(pathDatLog)
                    excelPrevTime = datetime.datetime.now()
                    logging.info("Excel File Saved")
                cursor12 = db.cursor()
                sql12 = "SELECT * FROM trip_settings"
                cursor12.execute(sql12)
                tripSetting = list(cursor12.fetchall()[0])
                allTripSetting = gatherTripSetting(tripSetting, paramNum) 
                cursorX = db.cursor()
                sqlX = "SELECT * FROM transformer_status"
                cursorX.execute(sqlX)
                allStat = cursorX.fetchall()[0][1:paramNum+1]
                logging.debug("Dtest01 " + str(len(allStat)))
                cursor11 = db.cursor()
                sql11 = "SELECT * FROM trip_status"
                cursor11.execute(sql11)
                allTripStat = list(cursor11.fetchall()[0][1:paramNum+1])
                logging.debug("Dtest02 " + str(len(allTripStat)))
                for j in range(0, paramNum):
                    if currentState[j] == 5 or currentState[j] == 1:
                        if allTripSetting[j] == 0:
                            allTripStat[j] = 1
                        elif allTripSetting[j] == 1:
                            allTripStat[j] = 2
                        elif allTripSetting[j] == 2:
                            allTripStat[j] = 3
                    else:
                        allTripStat[j] = 0
                    if previousState[j] != currentState[j]:
                        stateListBool[j] = 1
                        previousState[j] = currentState[j]
                        cursor5 = db.cursor()
                        sql5 = """UPDATE transformer_status SET 
                                Van = %s , Vbn = %s , Vcn = %s , Vab = %s , Vbc = %s , Vca = %s , 
                                Uab = %s , Ubc = %s , Uca = %s , Freq = %s , OilTemp = %s , 
                                WTITemp1 = %s , WTITemp2 = %s , WTITemp3 = %s ,
                                BusTemp1 = %s , BusTemp2 = %s , BusTemp3 = %s ,
                                PF = %s , Current1 = %s , Current2 = %s , Current3 = %s , 
                                AmbTemp = %s , Pressure = %s , OilLevel = %s, 
                                THDCurrent1 = %s, THDCurrent2 = %s, THDCurrent3 = %s,    
                                THDVoltage1 = %s, THDVoltage2 = %s, THDVoltage3 = %s, 
                                Ineutral = %s
                                WHERE trafoId = 1"""
                        cursor5.execute(sql5, currentState)
                        sekarang = datetime.datetime.now()
                        if j == 21 and gasFault:
                            logging.warning("Gas Fault Occurred")
                            cursor7 = db.cursor()
                            sql7 = "INSERT INTO failure_log (time_start, duration, failure_type, parameter) VALUES (%s, %s, %s, %s)"
                            errorVal = [sekarang, 0,  "Gas Fault", "Oil Level"]
                            cursor7.execute(sql7, errorVal)
                            errorMsg[j] = str("Oil Level Gas Fault !"  + "\n" + "Time Occurence : " + str(sekarang))
                        else:
                            if currentState[j] != 3:
                                if paramName[j] in activeParam:
                                    lastTimestamp = activeFailure[activeParam.index(paramName[j])][1]
                                    cursor9 = db.cursor()
                                    sql9 = "UPDATE failure_log SET duration = %s WHERE failure_id = %s"
                                    duration = int((sekarang - lastTimestamp).total_seconds())
                                    errorVal = [duration, activeFailure[activeParam.index(paramName[j])][0]]
                                    cursor9.execute(sql9, errorVal)
                                    activeFailure.pop(activeParam.index(paramName[j]))
                                    activeParam.pop(activeParam.index(paramName[j]))
                                cursor7 = db.cursor()
                                sql7 = "INSERT INTO failure_log (time_start, failure_type, parameter, parameterValue) VALUES (%s, %s, %s, %s)"
                                errorVal = [sekarang, messageReason[currentState[j]-1], paramName[j], str(paramValue[j])]
                                cursor7.execute(sql7, errorVal)
                                cursor8 = db.cursor()
                                sql8 = "SELECT * FROM failure_log ORDER BY failure_id DESC LIMIT 1"
                                cursor8.execute(sql8)
                                lastActive = cursor8.fetchall()[0]
                                activeFailure.append(lastActive)
                                errorMsg[j] = str(paramName[j] + " " + messageReason[previousState[j]-1] + " , Value = " + str(paramPrint[j]) + "\n" + "Time Occurence : " + str(sekarang))    
                            elif currentState[j] == 3:
                                try:
                                    lastTimestamp = activeFailure[activeParam.index(paramName[j])][1]
                                    cursor10 = db.cursor()
                                    sql10 = "UPDATE failure_log SET duration = %s WHERE failure_id = %s"
                                    duration = int((sekarang - lastTimestamp).total_seconds())
                                    errorVal = [duration, activeFailure[activeParam.index(paramName[j])][0]]
                                    cursor10.execute(sql10, errorVal)
                                    activeFailure.pop(activeParam.index(paramName[j]))
                                    activeParam.pop(activeParam.index(paramName[j]))
                                    errorMsg[j] = None
                                except Exception as Argument:
                                    logging.exception("Error returning faulted Value to normal")
                    else:
                        stateListBool[j] = 0
                    cursor13 = db.cursor()
                    sql13 = """UPDATE trip_status SET 
                            Van = %s , Vbn = %s , Vcn = %s , Vab = %s , Vbc = %s , Vca = %s , 
                            Uab = %s , Ubc = %s , Uca = %s , Freq = %s , OilTemp = %s , 
                            WTITemp1 = %s , WTITemp2 = %s , WTITemp3 = %s , 
                            BusTemp1 = %s , BusTemp2 = %s , BusTemp3 = %s , 
                            PF = %s , Current1 = %s , Current2 = %s , Current3 = %s , 
                            AmbTemp = %s , Pressure = %s , OilLevel = %s, 
                            THDCurrent1 = %s, THDCurrent2 = %s, THDCurrent3 = %s,    
                            THDVoltage1 = %s, THDVoltage2 = %s, THDVoltage3 = %s,
                            Ineutral = %s
                            WHERE trafoId = 1"""
                    cursor13.execute(sql13, allTripStat)          
            else:
                trialNum = trialNum + 1
                logging.warning("Retry Gathering Data")
            cursorX = db.cursor()
            sqlX = "SELECT * FROM transformer_status"
            cursorX.execute(sqlX)
            allStat = cursorX.fetchall()[0][1:paramNum+1]
            minStat = min(allStat)
            maxStat = max(allStat)
            cursor11 = db.cursor()
            sql11 = "SELECT * FROM trip_status"
            cursor11.execute(sql11)
            allTripStat = list(cursor11.fetchall()[0][1:paramNum+1])
            maxTripStat = max(allTripStat)
            actuatorStat = 0
            if minStat == 3 and maxStat == 3:
                actuatorStat = 0
                cursorY = db.cursor()
                sqlY = "UPDATE transformer_data SET status = 0 WHERE trafoId = 1"
                cursorY.execute(sqlY)
            elif minStat == 1 or maxStat == 5:
                if maxTripStat == 2:
                    actuatorStat = 3
                    cursorY = db.cursor()
                    sqlY = "UPDATE transformer_data SET status = 3 WHERE trafoId = 1"
                    cursorY.execute(sqlY)
                elif maxTripStat == 3:
                    actuatorStat = 2
                    cursorY = db.cursor()
                    sqlY = "UPDATE transformer_data SET status = 2 WHERE trafoId = 1"
                    cursorY.execute(sqlY)
                elif maxTripStat == 1:
                    actuatorStat = 1
                    cursorY = db.cursor()
                    sqlY = "UPDATE transformer_data SET status = 1 WHERE trafoId = 1"
                    cursorY.execute(sqlY)
                elif maxTripStat == 0:
                    actuatorStat = 0
                    cursorY = db.cursor()
                    sqlY = "UPDATE transformer_data SET status = 0 WHERE trafoId = 1"
            elif minStat == 2 or maxStat == 4:
                actuatorStat = 1
                cursorY = db.cursor()
                sqlY = "UPDATE transformer_data SET status = 1 WHERE trafoId = 1"
                cursorY.execute(sqlY)
            client.write_register(501, actuatorStat, slave=1)
            trafoStatLbl['text'] = str(actuatorStat)
            db.commit()
            stateChanges = any(stateListBool)
            errorMsgSend = list(filter(None, errorMsg))
            if stateChanges and any(errorMsgSend):
                logging.info("Sending Error Message to Telegram")
                sendTelegram = "\n-----------\n".join(map(str, errorMsgSend))
                messages = engineName + " Says : " + "\n" + sendTelegram
                pload = {'message':messages}
                try:
                    r = requests.post(teleURL, data = pload, timeout = 5, verify = False)
                except Timeout:
                    logging.warning("Timeout while sending tele message")
                except Exception as Argument:
                    logging.exception("Tele Catch")
                stateChanges = False
                stateListBool = [0]*len(stateListBool)
            logging.info("End Sect. Looping")
            time.sleep(5)
        else:
            progStatLbl['text'] = "Stop"
            
if __name__ == "__main__":
    restartBtn = Button(
        screen,
        text = "Restart",
        command = Restart)
    startBtn = Button(
        screen,
        text = "Start",
        command = Start)
    stopBtn = Button(
        screen,
        text = "Stop",
        command = Stop)
    lastTsLbl = Label(
            screen,
            font = ("Helvetica",9)
            )
    trafoStatLbl = Label(
            screen,
            font = ("Helvetica",9)
            )
    progStatLbl = Label(
            screen,
            font = ("Helvetica",9)
            )
    lastTsLbl.place(x = 10, y = 30)
    trafoStatLbl.place(x = 10, y = 70)
    progStatLbl.place(x = 10, y = 110)
    restartBtn.place(x = 315, y = 160)
    startBtn.place(x = 215, y = 160)
    stopBtn.place(x = 115, y = 160)
    thread1 = threading.Thread(target=mainLoop, args=('thread1', 1))
    thread1.start()
    screen.mainloop()
